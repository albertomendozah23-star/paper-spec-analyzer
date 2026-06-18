import pdfplumber
import re
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import io
import os


# ── helpers ──────────────────────────────────────────────────────────────────

def parse_float(s):
    if s is None:
        return None
    s = str(s).strip().replace(',', '.')
    m = re.search(r'-?\d+\.?\d*', s)
    return float(m.group()) if m else None


def parse_tolerance(tol_str):
    """Return (nominal, lsl, usl) from tolerance string like '70 +/- 3 %', '60 -10', '14 -2'"""
    if not tol_str:
        return None, None, None
    tol_str = str(tol_str).strip()

    # "+/- X" symmetric
    m = re.search(r'(\d+[\.,]?\d*)\s*\+/-\s*(\d+[\.,]?\d*)\s*(%?)', tol_str)
    if m:
        nom = float(m.group(1).replace(',', '.'))
        delta = float(m.group(2).replace(',', '.'))
        if m.group(3) == '%':
            delta = nom * delta / 100
        return nom, nom - delta, nom + delta

    # "NOM +UPPER / -LOWER"  e.g. "250 +150 / -100"
    m = re.search(r'(\d+[\.,]?\d*)\s*\+(\d+[\.,]?\d*)\s*/\s*-(\d+[\.,]?\d*)', tol_str)
    if m:
        nom = float(m.group(1).replace(',', '.'))
        return nom, nom - float(m.group(3).replace(',', '.')), nom + float(m.group(2).replace(',', '.'))

    # "NOM -X"  e.g. "60 -10"
    m = re.search(r'(\d+[\.,]?\d*)\s*-\s*(\d+[\.,]?\d*)$', tol_str)
    if m:
        nom = float(m.group(1).replace(',', '.'))
        delta = float(m.group(2).replace(',', '.'))
        return nom, nom - delta, nom

    # "X min." or "X max."
    m_min = re.search(r'(\d+[\.,]?\d*)\s*min', tol_str, re.I)
    m_max = re.search(r'(\d+[\.,]?\d*)\s*max', tol_str, re.I)
    if m_min:
        return float(m_min.group(1).replace(',', '.')), float(m_min.group(1).replace(',', '.')), None
    if m_max:
        return float(m_max.group(1).replace(',', '.')), None, float(m_max.group(1).replace(',', '.'))

    return None, None, None


def c4_factor(n):
    """Unbiasing constant c4 for sample size n (approximation valid for n>=2)."""
    if n is None or n < 2:
        return 1.0
    import math
    # c4 = sqrt(2/(n-1)) * gamma(n/2) / gamma((n-1)/2)
    try:
        return math.sqrt(2 / (n - 1)) * math.exp(
            math.lgamma(n / 2) - math.lgamma((n - 1) / 2)
        )
    except Exception:
        return 1.0


def compute_indices(value, std_dev, n, lsl, usl):
    """
    Compute Pp, Ppk (long-term, using s from PDF directly) and
    Cp, Cpk (short-term, using s/c4(n) as estimate of sigma).
    """
    results = {}
    if not (std_dev and std_dev > 0 and value is not None):
        return results

    s = std_dev  # long-term std dev (Pp/Ppk)
    c4 = c4_factor(n)
    sigma = s / c4  # short-term sigma estimate (Cp/Cpk)

    def _pp_ppk(sig):
        r = {}
        if lsl is not None and usl is not None:
            r['p'] = round((usl - lsl) / (6 * sig), 3)
            r['pk'] = round(min((usl - value) / (3 * sig), (value - lsl) / (3 * sig)), 3)
        elif usl is not None:
            r['pk'] = round((usl - value) / (3 * sig), 3)
        elif lsl is not None:
            r['pk'] = round((value - lsl) / (3 * sig), 3)
        return r

    pp_data = _pp_ppk(s)
    cp_data = _pp_ppk(sigma)

    results['Pp']  = pp_data.get('p')
    results['Ppk'] = pp_data.get('pk')
    results['Cp']  = cp_data.get('p')
    results['Cpk'] = cp_data.get('pk')
    return results


# ── extraction ────────────────────────────────────────────────────────────────

def extract_rows(pdf_path):
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue
                header = [str(c).lower().strip() if c else '' for c in table[0]]
                if 'property' not in ' '.join(header):
                    continue
                for raw in table[1:]:
                    if not raw or not raw[0]:
                        continue
                    prop_raw = str(raw[0]).replace('\n', ' ').strip()
                    unit = str(raw[1]).strip() if raw[1] else ''
                    method = str(raw[2]).replace('\n', ' ').strip() if raw[2] else ''
                    value_raw = str(raw[3]).replace('\n', '\n').strip() if raw[3] else ''
                    std_raw = str(raw[4]).strip() if len(raw) > 4 and raw[4] else ''
                    n_raw = str(raw[5]).strip() if len(raw) > 5 and raw[5] else ''
                    tol_raw = str(raw[6]).replace('\n', ' ').strip() if len(raw) > 6 and raw[6] else ''

                    # split multi-line sub-rows (CS/RS, MD/CD, PT/COT)
                    val_lines = value_raw.split('\n')
                    std_lines = std_raw.split('\n')
                    n_lines = n_raw.split('\n')
                    tol_lines = tol_raw.split('\n')
                    prop_lines = prop_raw.split('  ')  # Climatic Curl 30/50/80

                    n = max(len(val_lines), 1)
                    for i in range(n):
                        vl = val_lines[i].strip() if i < len(val_lines) else ''
                        sl = std_lines[i].strip() if i < len(std_lines) else (std_lines[0] if std_lines else '')
                        nl = n_lines[i].strip() if i < len(n_lines) else (n_lines[0] if n_lines else '')
                        tl = tol_lines[i].strip() if i < len(tol_lines) else (tol_lines[0] if tol_lines else '')

                        # property label
                        if n > 1:
                            # detect prefix like CS:, RS:, MD:, CD:, PT:, COT:
                            pm = re.match(r'^(CS|RS|MD|CD|PT|COT):\s*(.*)', vl)
                            if pm:
                                sub = pm.group(1)
                                vl = pm.group(2)
                                label = f"{prop_raw.split(chr(10))[0]} ({sub})"
                            else:
                                pl = prop_lines[i].strip() if i < len(prop_lines) else prop_raw.split('\n')[0]
                                label = pl if pl else prop_raw.split('\n')[0]
                        else:
                            label = prop_raw.split('\n')[0]

                        value = parse_float(vl)
                        std = parse_float(sl)
                        n_meas = parse_float(nl)
                        nom, lsl, usl = parse_tolerance(tl)

                        n_int = int(n_meas) if n_meas else None
                        indices = {}
                        if value is not None and std is not None:
                            indices = compute_indices(value, std, n_int, lsl, usl)

                        rows.append({
                            'property': label,
                            'unit': unit,
                            'method': method,
                            'value': value,
                            'std_dev': std,
                            'n': n_int,
                            'nominal': nom,
                            'lsl': lsl,
                            'usl': usl,
                            'tolerance_raw': tl,
                            'Pp':  indices.get('Pp'),
                            'Ppk': indices.get('Ppk'),
                            'Cp':  indices.get('Cp'),
                            'Cpk': indices.get('Cpk'),
                        })
    return rows


# ── chart helpers ─────────────────────────────────────────────────────────────

def cpk_color(cpk):
    if cpk is None:
        return '#888888'
    if cpk >= 1.33:
        return '#2ecc71'
    if cpk >= 1.0:
        return '#f39c12'
    return '#e74c3c'


def make_gauge_chart(label, value, std, lsl, usl, pp, ppk, cp, cpk):
    fig, axes = plt.subplots(1, 2, figsize=(9, 3.5), gridspec_kw={'width_ratios': [2, 1]})
    ax = axes[0]

    if lsl is not None or usl is not None:
        lo = (lsl - 4 * std) if lsl is not None else (value - 4 * std)
        hi = (usl + 4 * std) if usl is not None else (value + 4 * std)
        x = np.linspace(lo, hi, 400)
        y = (1 / (std * np.sqrt(2 * np.pi))) * np.exp(-0.5 * ((x - value) / std) ** 2)
        ax.plot(x, y, 'b-', lw=2)
        ax.fill_between(x, y, alpha=0.15, color='blue')

        if lsl is not None:
            ax.axvline(lsl, color='red', ls='--', lw=1.5, label=f'LSL={lsl}')
        if usl is not None:
            ax.axvline(usl, color='red', ls='--', lw=1.5, label=f'USL={usl}')
        ax.axvline(value, color='navy', ls='-', lw=2, label=f'Mean={value}')
        ax.set_xlabel('Value', fontsize=9)
        ax.set_ylabel('Density', fontsize=9)
        ax.legend(fontsize=8)
    else:
        ax.text(0.5, 0.5, 'No numeric\ntolerance', ha='center', va='center', transform=ax.transAxes, fontsize=11)

    ax.set_title(label, fontsize=10, fontweight='bold')
    ax.tick_params(labelsize=8)
    ax.grid(True, alpha=0.3)

    ax2 = axes[1]
    ax2.axis('off')
    lines = [
        ('Value',  f'{value}' if value is not None else 'N/A'),
        ('Std Dev', f'{std}'  if std   is not None else 'N/A'),
        ('LSL',    f'{lsl}'   if lsl   is not None else '—'),
        ('USL',    f'{usl}'   if usl   is not None else '—'),
        ('Pp',     f'{pp:.3f}'  if pp  is not None else 'N/A'),
        ('Ppk',    f'{ppk:.3f}' if ppk is not None else 'N/A'),
        ('Cp',     f'{cp:.3f}'  if cp  is not None else 'N/A'),
        ('Cpk',    f'{cpk:.3f}' if cpk is not None else 'N/A'),
    ]
    for j, (k, v) in enumerate(lines):
        is_ppk = k == 'Ppk' and ppk is not None
        is_cpk = k == 'Cpk' and cpk is not None
        color = cpk_color(ppk) if is_ppk else (cpk_color(cpk) if is_cpk else 'black')
        ax2.text(0.05, 0.97 - j * 0.12, f'{k}:', fontsize=9, va='top', fontweight='bold')
        ax2.text(0.55, 0.97 - j * 0.12, v, fontsize=9, va='top', color=color)

    fig.tight_layout(pad=1.5)
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=110, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf


def make_summary_chart(rows):
    valid = [(r['property'], r['Ppk'], r['Cpk']) for r in rows if r['Ppk'] is not None or r['Cpk'] is not None]
    if not valid:
        return None
    labels = [v[0] for v in valid]
    ppks   = [v[1] for v in valid]
    cpks   = [v[2] for v in valid]

    x = np.arange(len(labels))
    width = 0.35
    fig, ax = plt.subplots(figsize=(max(9, len(labels) * 1.1), 4.5))

    bars1 = ax.bar(x - width/2, ppks, width, label='Ppk (largo plazo)',
                   color=[cpk_color(v) for v in ppks], edgecolor='white', alpha=0.85)
    bars2 = ax.bar(x + width/2, cpks, width, label='Cpk (corto plazo)',
                   color=[cpk_color(v) for v in cpks], edgecolor='white', alpha=0.55, hatch='//')

    ax.axhline(1.33, color='green', ls='--', lw=1.5)
    ax.axhline(1.0,  color='orange', ls='--', lw=1.5)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=40, ha='right', fontsize=8)
    ax.set_ylabel('Índice de Capacidad')
    ax.set_title('Resumen de Capacidad de Proceso (Ppk y Cpk)', fontsize=12, fontweight='bold')
    ax.grid(True, axis='y', alpha=0.3)

    patches = [
        mpatches.Patch(color='#2ecc71', label='≥ 1.33 Capaz'),
        mpatches.Patch(color='#f39c12', label='1.00–1.33 Marginal'),
        mpatches.Patch(color='#e74c3c', label='< 1.00 No capaz'),
        mpatches.Patch(facecolor='white', edgecolor='grey', label='Ppk (sólido) / Cpk (rayado)', hatch='//'),
    ]
    ax.legend(handles=patches, fontsize=8, loc='upper right')
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=110, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf


# ── Excel builder ─────────────────────────────────────────────────────────────

GOLD = 'FFD700'
DARK = '2C2C2C'
LIGHT_GREY = 'F5F5F5'
GREEN = '2ecc71'
ORANGE = 'f39c12'
RED = 'e74c3c'

def hdr_style(cell, bg=GOLD):
    cell.font = Font(bold=True, color=DARK, size=10, name='Arial')
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='AAAAAA')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def data_style(cell, bold=False, color=None, bg=None):
    cell.font = Font(bold=bold, color=color or DARK, size=10, name='Arial')
    if bg:
        cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='DDDDDD')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_excel(rows, output_path, pdf_name=''):
    wb = Workbook()

    # ── Sheet 1: Data ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Data'
    ws.sheet_view.showGridLines = False

    # title
    ws.merge_cells('A1:L1')
    t = ws['A1']
    t.value = f'Paper Specification – {pdf_name}'
    t.font = Font(bold=True, size=14, color='FFFFFF', name='Arial')
    t.fill = PatternFill('solid', start_color='1A1A2E')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    headers = ['Property', 'Unit', 'Test Method', 'Value', 'Std Dev', 'N', 'Nominal', 'LSL', 'USL', 'Tolerance', 'Pp', 'Ppk', 'Cp', 'Cpk']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        hdr_style(c)

    ws.row_dimensions[2].height = 28

    for r_i, row in enumerate(rows, 3):
        vals = [
            row['property'], row['unit'], row['method'],
            row['value'], row['std_dev'], row['n'],
            row['nominal'], row['lsl'], row['usl'],
            row['tolerance_raw'], row['Pp'], row['Ppk'], row['Cp'], row['Cpk'],
        ]
        bg = LIGHT_GREY if r_i % 2 == 0 else 'FFFFFF'
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r_i, column=col, value=v)
            cpk_val = row['Cpk']
            if col == 14 and cpk_val is not None:  # Cpk column
                color = ('2ecc71' if cpk_val >= 1.33 else ('f39c12' if cpk_val >= 1.0 else 'e74c3c'))
                data_style(c, bold=True, color='FFFFFF', bg=color)
            elif col == 12 and row['Ppk'] is not None:  # Ppk column
                ppk_val = row['Ppk']
                color = ('2ecc71' if ppk_val >= 1.33 else ('f39c12' if ppk_val >= 1.0 else 'e74c3c'))
                data_style(c, bold=True, color='FFFFFF', bg=color)
            else:
                data_style(c, bg=bg)

    col_widths = [28, 10, 28, 10, 10, 6, 10, 10, 10, 20, 8, 8, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A3'

    # ── Sheet 2: Summary Chart ────────────────────────────────────────────────
    ws2 = wb.create_sheet('Capability Summary')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:J1')
    t2 = ws2['A1']
    t2.value = 'Process Capability Summary'
    t2.font = Font(bold=True, size=14, color='FFFFFF', name='Arial')
    t2.fill = PatternFill('solid', start_color='1A1A2E')
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 30

    sum_buf = make_summary_chart(rows)
    if sum_buf:
        img = XLImage(sum_buf)
        img.anchor = 'A3'
        img.width = 820
        img.height = 360
        ws2.add_image(img)

    # legend table below
    legend_row = 26
    for i, (label, color) in enumerate([
        ('Cpk ≥ 1.33', '2ecc71'), ('1.00 ≤ Cpk < 1.33', 'f39c12'), ('Cpk < 1.00', 'e74c3c')
    ]):
        c = ws2.cell(row=legend_row, column=i * 2 + 1, value=label)
        c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill = PatternFill('solid', start_color=color)
        c.alignment = Alignment(horizontal='center')

    # ── Sheet 3+: Individual capability charts ────────────────────────────────
    chart_rows = [r for r in rows if r['value'] is not None and r['std_dev'] is not None]
    ws3 = wb.create_sheet('Capability Charts')
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells('A1:N1')
    t3 = ws3['A1']
    t3.value = 'Individual Process Capability Charts'
    t3.font = Font(bold=True, size=14, color='FFFFFF', name='Arial')
    t3.fill = PatternFill('solid', start_color='1A1A2E')
    t3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 30

    cur_row = 3
    col_pos = 1
    per_row = 2

    for idx, row in enumerate(chart_rows):
        buf = make_gauge_chart(
            row['property'], row['value'], row['std_dev'],
            row['lsl'], row['usl'],
            row['Pp'], row['Ppk'], row['Cp'], row['Cpk']
        )
        img = XLImage(buf)
        img.width = 480
        img.height = 190
        col_letter = get_column_letter(col_pos)
        img.anchor = f'{col_letter}{cur_row}'
        ws3.add_image(img)

        col_pos += 7
        if (idx + 1) % per_row == 0:
            cur_row += 13
            col_pos = 1

    wb.save(output_path)
    return output_path


# ── public API ────────────────────────────────────────────────────────────────

def process_pdf(pdf_path, output_path):
    pdf_name = os.path.basename(pdf_path)
    rows = extract_rows(pdf_path)
    build_excel(rows, output_path, pdf_name)
    return output_path, len(rows)
